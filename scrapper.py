import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import time
import os
from colorama import Fore

start_time = time.time()  

def kirmizi_metin(text):
    return f"\033[38;2;255;0;0m{text}\033[0m"  

def yesil_metin(text):
    return f"\033[38;2;0;190;0m{text}\033[0m" 

def sari_metin(text):
    return f"\033[38;2;255;255;0m{text}\033[0m"

os.system('cls' if os.name == 'nt' else 'clear')
excel_file = "list.xlsx"

basarili_islem = 0

try:
    wb = load_workbook(excel_file)
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    header = ["Email", "Vacancy", "Employer Name", "Contact name","Mobile", "Company name", "Address", "Industry"]
    ws.append(header)
    fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    for cell in ws[1]:
        cell.fill = fill
    for col in ws.columns:
        max_length = 40
        column = col[0].column_letter  
        ws.column_dimensions[column].width = max_length

chrome_options = Options()
chrome_options.add_experimental_option('excludeSwitches', ["enable-automation", 'enable-logging'])
chrome_options.add_argument('--disable-logging')
chrome_options.add_argument('--log-level=3')
chrome_options.add_argument('--disable-infobars')
chrome_options.add_argument('--disable-extensions')
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument('--disable-gpu')

driver = webdriver.Chrome(options=chrome_options)
wait = WebDriverWait(driver, 6) 

links_df = pd.read_csv('links.csv', header=None, names=['link'])

for index, row in links_df.iterrows():
    url = row['link']
    driver.get(url)
    time.sleep(2)
    try:
        accept_cookies_button = driver.find_element(By.CSS_SELECTOR, 'a.wt-ecl-button.wt-ecl-button--primary.cck-actions-button')
        accept_cookies_button.click()
    except:
        pass
    
    
    try:
        email_address = ""
        phone_number = ""
        vacancy = ""
        job_location = ""
        employer_name = ""
        job_sector = ""
        company_name = ""
        contact_name = ""
        try:
            email_elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'a[href^="mailto:"]')))
            email_address = email_elem.text.strip()
        except:
            pass

        try:
            phone_elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'li#jv-details-telNumber-0-0')))
            
            phone_number_raw = phone_elem.text
            
            phone_number = phone_number_raw.replace('Tel.:', '').strip()
        except:
            pass


        try:
            vacancy_spans = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'span[id^="jv-job-categories-codes-result-"]')))
            vacancy = ' - '.join([span.text.strip() for span in vacancy_spans])
        except:
            pass

        try:
            location_elem = wait.until(EC.presence_of_element_located((By.ID, "jv-details-job-location")))
            job_location = location_elem.text.strip()
        except:
            pass

        try:
            employer_name_elem = wait.until(EC.presence_of_element_located((By.ID, "jv-details-employer-name")))
            employer_name = employer_name_elem.text.strip()
        except:
            pass

        try:
            job_sector_elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'dd.ecl-description-list__definition span#jv-employer-sector-codes-result')))
            job_sector = job_sector_elem.text.strip()
        except:
            pass

        try: 
            company_name_elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'h2.ecl-u-type-heading-2.ecl-u-mt-s')))
            company_name = company_name_elem.text.strip()
        except:
            pass

        try:
            contact_name_elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'li#jv-details-displayName-0')))
            contact_name = contact_name_elem.text.strip()
        except:
            pass


      
        ws.append([
            email_address,
            vacancy,
            employer_name,
            contact_name,
            phone_number,
            company_name,
            job_location,
            job_sector,
        ])

        links_df.drop(index, inplace=True)

    except Exception as e:
        print(Fore.RED + f"Error processing {url}: {e}" + Fore.RESET)

    wb.save(excel_file)

    basarili_islem += 1
    
    print(Fore.YELLOW + f"[ + ] >> [ Web Sitesinden Veri Kazıldı! ] >> [ Toplam Başarılı İşlem Sayısı {basarili_islem} ]" + Fore.RESET)

links_df.to_csv('links.csv', index=False, header=False)

end_time = time.time()  # Bitiş zamanı
elapsed_time = end_time - start_time  # Geçen süre

input(Fore.GREEN + f"\n[ + ] >> [ Bütün Veriler Excel Dosyasına Kaydedildi! ] >> [ Toplam Başarılı İşlem Sayısı {basarili_islem}] >> [ Geçen Süre: {elapsed_time:.2f} Saniye ]" + Fore.RESET)

driver.quit()
